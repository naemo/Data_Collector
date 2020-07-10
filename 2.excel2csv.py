## portMIS_Crawler로 다운로드한 액셀파일 병합

import pandas as pd
from openpyxl import load_workbook
import os

columns = ['Harbor', 'Date', 'isKorean', 'Full_10', 'Empty_10','Full_20', 'Empty_20',
           'Full_40', 'Empty_40', 'Full_other', 'Empty_other', 'R/T']
harbor_code = {'020':'Busan', '030':'Incheon', '031':'Pyeongtaek, Dangjin', '050':'Gyeongin Port',
               '200':'East Sea, Mukho', '201':'Samcheok', '203':'Sokcho', '204':'Okgye', '206':'Hosan',
               '300':'Daesan', '301':'Boryeong', '302':'Taean', '500':'Gunsan', '501':'Janghang',
               '610':'Mokpo', '611':'Wando', '620':'Yeosu', '622':'Gwangyang', '700':'Pohang',
               '810':'Masan', '820':'Ulsan', '900':'Jeju', '901':'Seogwipo'}

def makeDates():
    dates = [0, ]
    for y in range(2018, 2021):
        for m in range(1, 13):
            dates.append(y * 100 + m)
            dates.append(y * 100 + m)
            if y==2020 and m==5:
                return dates
dates = makeDates()

DIR = 'C:\\Users\\nakag\\Downloads\\'

#경로의 액셀파일 불러오기
directory = os.listdir(DIR)
datas = []
for f in directory:
    if f[-5:] == '.xlsx':
        xlsx = load_workbook(DIR+f, data_only=True)
        datas.append((xlsx, f[:-5]))

def read_row(ex, row):
    f10 = ex.cell(row, 4).value
    e10 = ex.cell(row, 5).value
    f20 = ex.cell(row, 7).value
    e20 = ex.cell(row, 8).value
    f40 = ex.cell(row, 10).value
    e40 = ex.cell(row, 11).value
    fo = ex.cell(row, 13).value
    eo = ex.cell(row, 14).value
    rt = ex.cell(row, 19).value
    return [f10, e10, f20, e20, f40, e40, fo, eo, rt]

rows = []
for ex, code in datas:
    print(code)
    sheet = ex.active #활성화된 시트 불러오기
    adder = 4
    for r in range(1, 59):
        row = []
        row.append(harbor_code[code])
        date = dates[r]*100+1
        row.append(str(date))
        kor = True
        if r%2==0:
            kor = False
        row.append(kor)
        row.extend(read_row(sheet, r+adder))
        rows.append(row)
        if r%2==0:
            adder += 1
        if r%24==0:
            adder += 1
    print(harbor_code[code], 'complete')

res = pd.DataFrame(rows, columns=columns)
res["Date"] = res["Date"].astype('datetime64[ns]')
res.to_csv('harbor.csv')